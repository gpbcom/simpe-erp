from __future__ import annotations

# Standard library imports
from decimal import Decimal
from io import BytesIO
from logging import getLogger
from typing import ClassVar, Dict, List, Optional, Tuple

# Third-party imports
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# First-party imports
from models.companies.company import Company
from models.enums import Language
from models.people.customer import Customer
from models.planning.hca_planning import HcaPlanning
from models.quoting.quote import Quote
from models.quoting.quote_line import QuoteLine


class Formatter:
    """Renders a quote or a planning as an Excel workbook.

    Attributes:
        PLANNING_HEADERS (ClassVar[Tuple[str, ...]]): Column titles of a
            planning sheet.
        QUOTE_HEADERS (ClassVar[Tuple[str, ...]]): Column titles of a quote's
            line sheet.
        HEADER_FILL (ClassVar[str]): Background colour of the header row.
        DATE_FORMAT (ClassVar[str]): How a day is written in a cell.
        MONEY_FORMAT (ClassVar[str]): How an amount is written in a cell.

    Notes:
        - Every method is static, and deliberately: formatting is a pure
          function of what it is handed. There is nothing to configure, nothing
          to inject and nothing to remember between two calls, so an instance
          would carry no state and only invite one to be added later.
        - Excel rather than CSV because both documents are read by people, not
          parsed: an assistant reads their week on a phone, a customer reads
          amounts they are being asked to pay. Column widths, a frozen header
          and real date and currency cells are the difference between a
          document and a dump.
        - Amounts are written as **numbers with a currency format**, never as
          pre-formatted strings. A string cannot be summed, and the first
          thing anyone does with a quote is add up a column.
    """

    PLANNING_HEADERS: ClassVar[Tuple[str, ...]] = (
        "Day",
        "Start",
        "End",
        "Service",
        "Customer",
        "Address",
        "Status",
    )
    QUOTE_HEADERS: ClassVar[Dict[Language, Tuple[str, ...]]] = {
        Language.EN: (
            "Date",
            "Service",
            "Quantity (h)",
            "Unit price (excl. VAT)",
            "VAT rate",
            "Amount (excl. VAT)",
            "VAT",
            "Amount (incl. VAT)",
        ),
        Language.FR: (
            "Date",
            "Prestation",
            "Quantité (h)",
            "Prix unitaire HT",
            "Taux TVA",
            "Montant HT",
            "TVA",
            "Montant TTC",
        ),
    }
    QUOTE_LABELS: ClassVar[Dict[Language, Dict[str, str]]] = {
        Language.EN: {
            "title": "Quote",
            "issued_by": "Issued by",
            "customer": "Customer",
            "status": "Status",
            "valid_until": "Valid until",
            "period": "Period",
            "interventions": "Interventions",
            "total_duration": "Total duration",
            "total_incl_vat": "Total (incl. VAT)",
            "total": "Total",
            "registration": "Reg. no.",
            "vat_number": "VAT",
            "share_capital": "Share capital",
            "sheet": "Quote",
            "no_line": "This quote has no service line.",
            "status_draft": "draft",
            "status_pending-validation": "pending validation",
            "status_sent": "sent",
            "status_accepted": "accepted",
            "status_rejected": "rejected",
            "status_expired": "expired",
            "untaxed_amount": "Untaxed Amount",
            "tax_at": "Tax {rate}",
            "grand_total": "Total",
            "legal_heading": "Terms",
            "validity": "This quote is valid until {date}.",
            "payment_terms": (
                "Payment is due within 30 days of the date of the invoice. "
                "Please quote {reference} when paying."
            ),
            "late_penalty": (
                "Late payment carries interest at three times the statutory "
                "rate, plus a fixed 40 € recovery charge (French Commercial "
                "Code, art. L441-10)."
            ),
            "acceptance": (
                "Quote received before the work begins. Return signed, dated "
                "and marked \u00ab Bon pour accord \u00bb to accept."
            ),
            "signature": "Date and signature of the customer:",
        },
        Language.FR: {
            "title": "Devis",
            "issued_by": "Émis par",
            "customer": "Client",
            "status": "Statut",
            "valid_until": "Valable jusqu'au",
            "period": "Période",
            "interventions": "Interventions",
            "total_duration": "Durée totale",
            "total_incl_vat": "Total TTC",
            "total": "Total",
            "registration": "SIRET",
            "vat_number": "TVA",
            "share_capital": "Capital social",
            "sheet": "Devis",
            "no_line": "Ce devis ne comporte aucune prestation.",
            "status_draft": "brouillon",
            "status_pending-validation": "en attente de validation",
            "status_sent": "envoyé",
            "status_accepted": "accepté",
            "status_rejected": "refusé",
            "status_expired": "expiré",
            "untaxed_amount": "Total HT",
            "tax_at": "TVA {rate}",
            "grand_total": "Total TTC",
            "legal_heading": "Conditions",
            "validity": "Le présent devis est valable jusqu'au {date}.",
            "payment_terms": (
                "Paiement à 30 jours à compter de la date de facture. "
                "Merci de rappeler la référence {reference} lors du règlement."
            ),
            "late_penalty": (
                "Tout retard de paiement entraîne des pénalités au taux légal "
                "majoré de trois fois, ainsi qu'une indemnité forfaitaire de "
                "recouvrement de 40 € (code de commerce, art. L441-10)."
            ),
            "acceptance": (
                "Devis reçu avant l'exécution des travaux. À retourner signé "
                "et daté avec la mention \u00ab Bon pour accord \u00bb."
            ),
            "signature": "Date et signature du client :",
        },
    }
    #: Where each block of the quote sheet begins. Named rather than written
    #: as literals at twenty call sites: the layout has grown twice, and each
    #: time every row below the insertion point had to move by hand.
    QUOTE_TITLE_ROW: ClassVar[int] = 1
    QUOTE_ISSUER_ROW: ClassVar[int] = 2
    QUOTE_CUSTOMER_ROW: ClassVar[int] = 5
    QUOTE_STATUS_ROW: ClassVar[int] = 7
    QUOTE_SUMMARY_ROW: ClassVar[int] = 9
    QUOTE_HEADER_ROW: ClassVar[int] = 12
    #: How large the agency's logo may be drawn, in pixels, and how tall the
    #: issuer rows are made so it fits between them and the customer block.
    LOGO_MAX_WIDTH: ClassVar[int] = 200
    LOGO_MAX_HEIGHT: ClassVar[int] = 56
    LOGO_ROW_HEIGHT: ClassVar[int] = 22
    #: How many columns from the right edge the logo is anchored.
    LOGO_COLUMN_INSET: ClassVar[int] = 2
    TITLE_FILL: ClassVar[str] = "FF1F3864"
    HEADER_FILL: ClassVar[str] = "FF2F5597"
    SUBTITLE_FILL: ClassVar[str] = "FFEDF2F9"
    BAND_FILL: ClassVar[str] = "FFEEF3FA"
    TOTAL_FILL: ClassVar[str] = "FFDDE5F0"
    GRID_COLOUR: ClassVar[str] = "FFD6DCE4"
    MUTED_TEXT: ClassVar[str] = "FF4A5568"
    STATUS_COLOURS: ClassVar[Dict[str, Tuple[str, str]]] = {
        "planned": ("FFDDEBF7", "FF1F4E79"),
        "confirmed": ("FFE2EFDA", "FF375623"),
        "completed": ("FFEDEDED", "FF3F3F3F"),
        "cancelled": ("FFFCE4E4", "FF9C0006"),
        "draft": ("FFEDEDED", "FF3F3F3F"),
        "sent": ("FFDDEBF7", "FF1F4E79"),
        "accepted": ("FFE2EFDA", "FF375623"),
        "rejected": ("FFFCE4E4", "FF9C0006"),
        "expired": ("FFFFF2CC", "FF7F6000"),
    }
    MAX_SHEET_TITLE: ClassVar[int] = 31
    FORBIDDEN_SHEET_CHARACTERS: ClassVar[Tuple[str, ...]] = (
        "/",
        "\\",
        "?",
        "*",
        "[",
        "]",
        ":",
    )
    DATE_FORMAT: ClassVar[str] = "yyyy-mm-dd"
    MONEY_FORMAT: ClassVar[str] = '#,##0.00\\ "€"'
    QUANTITY_FORMAT: ClassVar[str] = "#,##0.00"
    RATE_FORMAT: ClassVar[str] = "0.0%"

    @staticmethod
    def write_planning_sheet(sheet: Worksheet, planning: HcaPlanning) -> None:
        """Lay one assistant's week out on a worksheet.

        Args:
            sheet (Worksheet): The sheet to write on.
            planning (HcaPlanning): The week to render.

        Notes:
            - Shared by both documents, so an assistant's own sheet and their
              page of the team workbook are the same sheet. Two renderers would
              drift, and the first person to notice would be the assistant
              holding the one that is wrong.
            - Interventions are written in the order they happen, not the order
              they were solved in: this is the sheet an assistant works from,
              and a day out of sequence is worse than no sheet at all.
        """
        iso_year, iso_week, _ = planning.period_start.isocalendar()
        columns = len(Formatter.PLANNING_HEADERS)
        last_column = sheet.cell(row=4, column=columns).column_letter

        sheet.merge_cells(f"A1:{last_column}1")
        title = sheet["A1"]
        title.value = f"Planning — {planning.hca_full_name}"
        title.font = Font(bold=True, size=16, color="FFFFFFFF")
        title.fill = PatternFill("solid", fgColor=Formatter.TITLE_FILL)
        title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sheet.row_dimensions[1].height = 30

        sheet.merge_cells(f"A2:{last_column}2")
        subtitle = sheet["A2"]
        subtitle.value = (
            f"Week {iso_week:02d} of {iso_year} — "
            f"{planning.period_start} to {planning.period_end}"
        )
        subtitle.font = Font(bold=True, color=Formatter.MUTED_TEXT)
        subtitle.fill = PatternFill("solid", fgColor=Formatter.SUBTITLE_FILL)
        subtitle.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        sheet.merge_cells(f"A3:{last_column}3")
        summary = sheet["A3"]
        summary.value = (
            f"{len(planning.interventions)} intervention(s), "
            f"{planning.total_minutes() // 60}h{planning.total_minutes() % 60:02d}"
        )
        summary.font = Font(italic=True, color=Formatter.MUTED_TEXT)
        summary.fill = PatternFill("solid", fgColor=Formatter.SUBTITLE_FILL)
        summary.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        edge = Side(style="thin", color=Formatter.GRID_COLOUR)
        grid = Border(left=edge, right=edge, top=edge, bottom=edge)
        for column, header in enumerate(Formatter.PLANNING_HEADERS, start=1):
            cell = sheet.cell(row=4, column=column, value=header)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor=Formatter.HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = grid
        sheet.row_dimensions[4].height = 22

        ordered = sorted(
            planning.interventions, key=lambda item: (item.day, item.start_time)
        )
        band = PatternFill("solid", fgColor=Formatter.BAND_FILL)
        for offset, intervention in enumerate(ordered):
            row = 5 + offset
            sheet.cell(
                row=row, column=1, value=intervention.day
            ).number_format = Formatter.DATE_FORMAT
            sheet.cell(row=row, column=2, value=intervention.start_time.isoformat())
            sheet.cell(row=row, column=3, value=intervention.end_time.isoformat())
            sheet.cell(row=row, column=4, value=intervention.name)
            sheet.cell(row=row, column=5, value=intervention.customer_id)
            sheet.cell(row=row, column=6, value=intervention.address.to_single_line())
            sheet.cell(row=row, column=7, value=intervention.status.value)

            background, text = Formatter.STATUS_COLOURS.get(
                intervention.status.value, (Formatter.BAND_FILL, Formatter.MUTED_TEXT)
            )
            for column in range(1, columns + 1):
                cell = sheet.cell(row=row, column=column)
                cell.border = grid
                if offset % 2:
                    cell.fill = band
                if column in (1, 2, 3):
                    cell.alignment = Alignment(horizontal="center")
            status_cell = sheet.cell(row=row, column=columns)
            status_cell.fill = PatternFill("solid", fgColor=background)
            status_cell.font = Font(bold=True, color=text)
            status_cell.alignment = Alignment(horizontal="center")

        if not ordered:
            empty = sheet.cell(row=5, column=1, value="No intervention scheduled.")
            empty.font = Font(italic=True, color=Formatter.MUTED_TEXT)
            sheet.merge_cells(f"A5:{last_column}5")
            getLogger(__name__).warning(
                "Assistant %s has an empty planning for %s to %s.",
                planning.hca_id,
                planning.period_start,
                planning.period_end,
            )
        else:
            total_row = 5 + len(ordered)
            label = sheet.cell(row=total_row, column=1, value="Total")
            label.font = Font(bold=True)
            hours = sheet.cell(
                row=total_row,
                column=2,
                value=(
                    f"{planning.total_minutes() // 60}h"
                    f"{planning.total_minutes() % 60:02d}"
                ),
            )
            hours.font = Font(bold=True)
            hours.alignment = Alignment(horizontal="center")
            for column in range(1, columns + 1):
                cell = sheet.cell(row=total_row, column=column)
                cell.fill = PatternFill("solid", fgColor=Formatter.TOTAL_FILL)
                cell.border = Border(
                    left=edge,
                    right=edge,
                    bottom=edge,
                    top=Side(style="double", color=Formatter.HEADER_FILL),
                )
            sheet.auto_filter.ref = f"A4:{last_column}{4 + len(ordered)}"

        widths: List[int] = [12, 8, 8, 34, 20, 52, 12]
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[
                sheet.cell(row=4, column=column).column_letter
            ].width = width
        sheet.freeze_panes = "A5"
        sheet.sheet_properties.tabColor = Formatter.TITLE_FILL
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_title_rows = "1:4"

    @staticmethod
    def format_planning(planning: HcaPlanning) -> bytes:
        """Render one assistant's own week as an Excel workbook.

        Args:
            planning (HcaPlanning): The week to render.

        Returns:
            bytes: The ``.xlsx`` document, ready to attach to an email.

        Notes:
            One sheet, one assistant, one week. This is what an assistant
            receives, and it contains nobody else's visits — the model itself
            only holds one person's, so there is nothing here to filter.
        """
        logger = getLogger(__name__)
        logger.debug(
            "Rendering the week of assistant %s (%s to %s, %d intervention(s)).",
            planning.hca_id,
            planning.period_start,
            planning.period_end,
            len(planning.interventions),
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Planning"
        Formatter.write_planning_sheet(sheet, planning)

        stream = BytesIO()
        workbook.save(stream)
        payload = stream.getvalue()
        logger.info(
            "Rendered a %d-byte planning workbook for assistant %s.",
            len(payload),
            planning.hca_id,
        )
        return payload

    @staticmethod
    def format_plannings(plannings: List[HcaPlanning]) -> bytes:
        """Render the whole workforce's week, one sheet per assistant.

        Args:
            plannings (List[HcaPlanning]): Every assistant's week.

        Returns:
            bytes: The ``.xlsx`` document, ready to attach to an email.

        Notes:
            - This is the manager's copy. A sheet per assistant rather than one
              long sheet with a name column: a manager reads one round at a
              time, and Excel's sheet tabs are the fastest index there is.
            - Sheet names are the assistant's name, trimmed to the 31 characters
              Excel allows and stripped of the five characters it forbids. Two
              assistants who collide after trimming are disambiguated by a
              counter rather than silently overwriting one another.
            - Assistants are ordered by name, so the same workbook two weeks
              running has its tabs in the same places.
        """
        logger = getLogger(__name__)
        logger.info(
            "Rendering a team planning workbook for %d assistant(s).", len(plannings)
        )
        workbook = Workbook()
        workbook.remove(workbook.active)

        ordered = sorted(plannings, key=lambda item: (item.hca_full_name, item.hca_id))
        used: List[str] = []
        for planning in ordered:
            title = planning.hca_full_name
            for forbidden in Formatter.FORBIDDEN_SHEET_CHARACTERS:
                title = title.replace(forbidden, " ")
            title = title.strip()[: Formatter.MAX_SHEET_TITLE] or planning.hca_id
            if title in used:
                suffix = f" ({used.count(title) + 1})"
                title = title[: Formatter.MAX_SHEET_TITLE - len(suffix)] + suffix
            used.append(title)
            Formatter.write_planning_sheet(workbook.create_sheet(title), planning)

        if not ordered:
            sheet = workbook.create_sheet("Planning")
            sheet["A1"] = "No assistant has a planning for this week."
            logger.warning("Rendered a team workbook with no assistant in it.")

        stream = BytesIO()
        workbook.save(stream)
        payload = stream.getvalue()
        logger.info(
            "Rendered a %d-byte team workbook holding %d sheet(s).",
            len(payload),
            len(workbook.sheetnames),
        )
        return payload

    @staticmethod
    def format_rate(rate: Decimal) -> str:
        """Render a VAT rate as a percentage a reader recognises.

        Args:
            rate (Decimal): The rate as a fraction, such as ``Decimal("0.055")``.

        Returns:
            str: ``"5.5%"`` or ``"20%"``.

        Notes:
            Trailing zeros are dropped. ``Decimal("0.055") * 100`` is
            ``Decimal("5.500")``, and a quote stating "TVA 5.500%" reads as a
            rate somebody mistyped. ``normalize`` is avoided because it renders
            twenty as ``2E+1``.
        """
        percent = rate * Decimal("100")
        text = f"{percent:f}".rstrip("0").rstrip(".")
        return f"{text}%"

    @staticmethod
    def write_quote_totals(
        sheet: Worksheet,
        quote: Quote,
        lines: List[QuoteLine],
        first_row: int,
        columns: int,
        labels: Dict[str, str],
    ) -> int:
        """Write the untaxed total, the tax per rate, and the grand total.

        Args:
            sheet (Worksheet): The sheet being built.
            quote (Quote): The quote being rendered.
            lines (List[QuoteLine]): The lines printed above, which the
                totals are summed from.
            first_row (int): The row just after the last line.
            columns (int): How wide the table is.
            labels (Dict[str, str]): The label catalogue for the language.

        Returns:
            int: The row the grand total was written on.

        Notes:
            - **The tax is stated per rate, and that is a legal requirement
              rather than a courtesy.** Home care is billed at 5.5% for a
              necessity service and 20% for a comfort one; a single "VAT" line
              gives the customer no way to check either figure and an accountant
              no way to post it.
            - Laid out as a right-hand stack rather than as a row under the
              table, which is what a reader expects of a document that ends in an
              amount — and what makes the grand total the last thing on the page
              rather than one cell among nine.
            - **Summed from the lines printed above, not from the quote's
              aggregates.** The two agree on a priced quote, but the aggregates
              are computed by the pricing service and a quote that reached this
              renderer without them would print a total of zero under a column
              of real amounts. On a document a customer is asked to sign, a
              total that disagrees with its own column is the worst failure
              available — worse than refusing to render at all, because it
              looks correct.
        """
        label_start = columns - 2
        row = first_row
        untaxed = sum(
            (line.total_ht or Decimal("0.00") for line in lines),
            Decimal("0.00"),  # noqa: E501
        )
        inclusive = sum(
            (line.total_ttc or Decimal("0.00") for line in lines),
            Decimal("0.00"),  # noqa: E501
        )
        entries = [(labels["untaxed_amount"], float(untaxed), False)]
        for rate, _, tax in quote.vat_by_rate():
            entries.append(
                (
                    labels["tax_at"].format(rate=Formatter.format_rate(rate)),
                    float(tax),
                    False,
                )
            )
        entries.append((labels["grand_total"], float(inclusive), True))

        for label, amount, emphasised in entries:
            row += 1
            sheet.merge_cells(
                start_row=row,
                start_column=label_start,
                end_row=row,
                end_column=columns - 1,
            )
            name = sheet.cell(row=row, column=label_start, value=label)
            name.alignment = Alignment(horizontal="right", indent=1)
            name.font = Font(bold=True, size=12 if emphasised else 11)
            value = sheet.cell(row=row, column=columns, value=amount)
            value.number_format = Formatter.MONEY_FORMAT
            value.alignment = Alignment(horizontal="right")
            value.font = Font(bold=True, size=12 if emphasised else 11)
            for column in range(label_start, columns + 1):
                cell = sheet.cell(row=row, column=column)
                if emphasised:
                    cell.fill = PatternFill("solid", fgColor=Formatter.TOTAL_FILL)
                    cell.border = Border(
                        top=Side(style="double", color=Formatter.HEADER_FILL),
                        bottom=Side(style="thin", color=Formatter.GRID_COLOUR),
                    )
                else:
                    cell.border = Border(
                        bottom=Side(style="thin", color=Formatter.GRID_COLOUR)
                    )
        return row

    @staticmethod
    def write_legal_terms(
        sheet: Worksheet,
        quote: Quote,
        first_row: int,
        columns: int,
        labels: Dict[str, str],
    ) -> None:
        """Write the mentions a quote must legally carry.

        Args:
            sheet (Worksheet): The sheet being built.
            quote (Quote): The quote being rendered.
            first_row (int): Where the block starts.
            columns (int): How wide the table is.
            labels (Dict[str, str]): The label catalogue for the language.

        Notes:
            - **A quote without these is not a quote, it is a price list.** French
              law requires a commercial offer to state how long it stands, when
              payment falls due, what happens if it does not, and to be returned
              signed before the work begins — and the customer's signature block
              is what turns the document into the agreement it claims to be.
            - The validity sentence is omitted when the quote carries no
              ``valid_until``. A draft has none yet, and an offer that promised to
              stand "until None" would be worse than one that promised nothing.
            - The wording is translated, but the obligations are French whatever
              language the reader chose: the agency is French and so is the
              contract. → the language decides the words, not the law.
        """
        row = first_row
        heading = sheet.cell(row=row, column=1, value=labels["legal_heading"])
        heading.font = Font(bold=True, color=Formatter.MUTED_TEXT)

        sentences = []
        if quote.valid_until is not None:
            sentences.append(labels["validity"].format(date=quote.valid_until))
        sentences.append(labels["payment_terms"].format(reference=quote.reference))
        sentences.append(labels["late_penalty"])
        sentences.append(labels["acceptance"])

        for sentence in sentences:
            row += 1
            sheet.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=columns
            )
            cell = sheet.cell(row=row, column=1, value=sentence)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            cell.font = Font(size=9, color=Formatter.MUTED_TEXT)

        row += 2
        sheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=columns
        )
        signature = sheet.cell(row=row, column=1, value=labels["signature"])
        signature.font = Font(bold=True, color=Formatter.MUTED_TEXT)
        sheet.row_dimensions[row].height = 46

    @staticmethod
    def anchor_logo(sheet: Worksheet, logo: bytes, anchor: str) -> bool:
        """Draw the agency's logo on the quote, scaled to fit its corner.

        Args:
            sheet (Worksheet): The sheet being built.
            logo (bytes): The image bytes, already fetched by the caller.
            anchor (str): The cell the image's top-left corner sits on.

        Returns:
            bool: ``True`` when the logo was drawn, ``False`` when the bytes
            could not be read as an image.

        Notes:
            - **Reports rather than raises.** A quote is a document a customer is
              waiting for, and refusing to produce one because its letterhead is
              corrupt would be the wrong trade — the caller carries on and the
              document goes out looking as it did before logos existed.
            - Scaled down but never up. Enlarging a small logo to fill the space
              would print it blurred, and an agency that uploaded a 60-pixel mark
              meant it to be small.
            - The issuer rows are made taller so the image sits inside them
              rather than over the customer's address. Openpyxl images float
              above the grid and do not push anything down.
        """
        logger = getLogger(__name__)
        logger.debug("Anchoring a %d-byte logo at %s.", len(logo), anchor)
        if not logo:
            logger.warning("The agency logo is empty; omitting it from the quote.")
            return False
        try:
            image = Image(BytesIO(logo))
        except Exception as exc:  # noqa: BLE001 - a bad logo must not stop a quote
            logger.error(
                "Could not read the agency logo, omitting it: %s. The quote is "
                "produced without a letterhead rather than not at all.",
                exc,
            )
            return False
        scale = min(
            Formatter.LOGO_MAX_WIDTH / image.width,
            Formatter.LOGO_MAX_HEIGHT / image.height,
            1,
        )
        image.width = int(image.width * scale)
        image.height = int(image.height * scale)
        image.anchor = anchor
        sheet.add_image(image)
        for row in (Formatter.QUOTE_ISSUER_ROW, Formatter.QUOTE_ISSUER_ROW + 1):
            sheet.row_dimensions[row].height = Formatter.LOGO_ROW_HEIGHT
        logger.info("Drew a %dx%d logo at %s.", image.width, image.height, anchor)
        return True

    @staticmethod
    def trading_name(company: Company) -> str:
        """Return the agency's name followed by its legal form.

        Args:
            company (Company): The agency issuing the quote.

        Returns:
            str: ``"Aide et Presence Paris SARL"``, or the bare name when no
            legal form is recorded.

        Notes:
            - The legal form is part of how a French company must identify itself
              on a commercial document — "Aide et Presence Paris" and "Aide et
              Presence Paris SARL" are a trading name and a legal person, and
              only the second is who the customer is contracting with.
            - Appended rather than required. An agency that has not filled it in
              yet prints its name alone, which is what the document said before
              this field existed; refusing to render would take the quote away
              from an agency over a field it can fill in afterwards.
        """
        if not company.legal_form:
            return company.name
        return f"{company.name} {company.legal_form}"

    @staticmethod
    def describe_company(company: Company, labels: Dict[str, str]) -> str:
        """Render the issuing agency as one line of contact detail.

        Args:
            company (Company): The agency issuing the quote.
            labels (Dict[str, str]): The label catalogue for the language.

        Returns:
            str: Address, registration number and contact address, whichever of
            them exist, separated by a middle dot.

        Notes:
            Built by joining only what is set. An agency with no registered
            number is ordinary — the field is optional — and printing
            "SIRET: None" beside a real address would read as a data error on a
            document a customer is being asked to sign.
        """
        parts = []
        if company.address is not None:
            parts.append(company.address.to_single_line())
        if company.registration_number:
            parts.append(f"{labels['registration']} {company.registration_number}")
        if company.rcs_number:
            parts.append(company.rcs_number)
        if company.vat_number:
            parts.append(f"{labels['vat_number']} {company.vat_number}")
        if company.share_capital is not None:
            parts.append(
                f"{labels['share_capital']} {company.share_capital:,.2f} €".replace(
                    ",", " "
                )
            )
        if company.phone_number:
            parts.append(company.phone_number)
        if company.contact_email:
            parts.append(str(company.contact_email))
        return " · ".join(parts)

    @staticmethod
    def write_quote_summary(
        sheet: Worksheet,
        quote: Quote,
        lines: List[QuoteLine],
        labels: Dict[str, str],
        language: Language,
    ) -> None:
        """Write the period, the count and the total duration above the table.

        Args:
            sheet (Worksheet): The sheet being built.
            quote (Quote): The quote being rendered.
            lines (List[QuoteLine]): Its lines, already ordered by date.
            labels (Dict[str, str]): The label catalogue for the language.
            language (Language): The language to write the duration in.

        Notes:
            - Four figures, laid out as label/value pairs rather than as a
              sentence: these are the questions asked of a quote *before* its
              price — when does this start, when does it end, how many visits is
              it, and how much care in total — and a reader scanning for one of
              them should not have to read the other three.
            - The period comes from the lines rather than from ``issued_on`` and
              ``valid_until``. Those two describe the *offer*; a customer asking
              "when does the care run from and to" is asking about the work.
        """
        row = Formatter.QUOTE_SUMMARY_ROW
        first = lines[0].service_date if lines else None
        last = lines[-1].service_date if lines else None
        period = f"{first} → {last}" if first and last else labels["no_line"]
        duration = Formatter.format_duration(
            sum(line.duration_minutes for line in lines), language
        )
        pairs = (
            (row, labels["period"], period),
            (row, labels["interventions"], len(lines)),
            (row + 1, labels["total_duration"], duration),
            (
                row + 1,
                labels["total_incl_vat"],
                float(
                    sum(
                        (line.total_ttc or Decimal("0.00") for line in lines),
                        Decimal("0.00"),
                    )
                ),
            ),
        )
        for index, (target, label, value) in enumerate(pairs):
            label_column, value_start, value_end = (
                (1, 3, 5) if index % 2 == 0 else (6, 7, 9)
            )
            label_cell = sheet.cell(row=target, column=label_column, value=label)
            label_cell.font = Font(bold=True, color=Formatter.MUTED_TEXT)
            sheet.merge_cells(
                start_row=target,
                start_column=label_column,
                end_row=target,
                end_column=value_start - 1,
            )
            value_cell = sheet.cell(row=target, column=value_start, value=value)
            value_cell.alignment = Alignment(horizontal="left", indent=1)
            if label == labels["total_incl_vat"]:
                value_cell.number_format = Formatter.MONEY_FORMAT
                value_cell.font = Font(bold=True)
            sheet.merge_cells(
                start_row=target,
                start_column=value_start,
                end_row=target,
                end_column=value_end,
            )
            for column in range(label_column, value_end + 1):
                sheet.cell(row=target, column=column).fill = PatternFill(
                    "solid", fgColor=Formatter.SUBTITLE_FILL
                )

    @staticmethod
    def format_duration(minutes: int, language: Language) -> str:
        """Render a number of minutes as hours and minutes.

        Args:
            minutes (int): The duration to render.
            language (Language): The language to render it in.

        Returns:
            str: ``"18 h 30"`` in French, ``"18h 30m"`` in English.

        Notes:
            Written out rather than left as a raw minute count. "1110" is the
            number the solver works in; a customer reading what they are being
            asked to pay for wants to know it is eighteen and a half hours.
        """
        hours, remainder = divmod(max(minutes, 0), 60)
        if language is Language.FR:
            return f"{hours} h {remainder:02d}"
        return f"{hours}h {remainder:02d}m"

    @staticmethod
    def format_quote(
        quote: Quote,
        customer: Customer,
        company: Company,
        language: Language = Language.FR,
        logo: Optional[bytes] = None,
    ) -> bytes:
        """Render a quote as an Excel workbook, in a given language.

        Args:
            quote (Quote): The quote to render, priced.
            customer (Customer): The customer it is addressed to.
            company (Company): The agency issuing it.
            language (Language): The language to write the document in.
            logo (Optional[bytes]): The agency's logo, already fetched, or
                ``None`` to print without one.

        Returns:
            bytes: The ``.xlsx`` document, ready to attach to an email.

        Notes:
            - **Both parties are named.** A quote carrying only the recipient is
              a document the recipient cannot act on: they need to know who is
              offering, at what registered number, and where to reply. The
              agency comes from the account that issued the quote — neither a
              customer nor a quote carries a company of its own.
            - The summary block states the period, the number of interventions
              and the **total duration**, because those are the three questions
              asked of a quote before its price: when does this start, when does
              it end, and how much care is it.
            - An unpriced line is written with empty amount cells rather than
              zeroes. A zero reads as "free"; an empty cell reads as "not priced
              yet", which is what it is.
            - The language is the one stored on the account that issued the
              quote. It is a stored preference rather than a request header
              because this runs in the planning-completed webhook, where there
              is no request.
        """
        logger = getLogger(__name__)
        labels = Formatter.QUOTE_LABELS[language]
        headers = Formatter.QUOTE_HEADERS[language]
        logger.debug(
            "Rendering quote %s for customer %s in %s (%d line(s)).",
            quote.reference,
            customer.id,
            language.value,
            len(quote.lines),
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = labels["sheet"]
        columns = len(headers)
        header_row = Formatter.QUOTE_HEADER_ROW
        last_column = sheet.cell(row=header_row, column=columns).column_letter
        edge = Side(style="thin", color=Formatter.GRID_COLOUR)
        grid = Border(left=edge, right=edge, top=edge, bottom=edge)

        sheet.merge_cells(f"A1:{last_column}1")
        title = sheet["A1"]
        title.value = f"{labels['title']} {quote.reference}"
        title.font = Font(bold=True, size=16, color="FFFFFFFF")
        title.fill = PatternFill("solid", fgColor=Formatter.TITLE_FILL)
        title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sheet.row_dimensions[1].height = 30

        issuer = Formatter.describe_company(company, labels)
        parties = (
            (
                Formatter.QUOTE_ISSUER_ROW,
                f"{labels['issued_by']}: {Formatter.trading_name(company)}",
            ),
            (Formatter.QUOTE_ISSUER_ROW + 1, issuer),
            (
                Formatter.QUOTE_CUSTOMER_ROW,
                f"{labels['customer']}: {customer.full_name()}",
            ),
            (Formatter.QUOTE_CUSTOMER_ROW + 1, customer.address.to_single_line()),
        )
        for row, value in parties:
            sheet.merge_cells(f"A{row}:{last_column}{row}")
            cell = sheet.cell(row=row, column=1, value=value)
            cell.fill = PatternFill("solid", fgColor=Formatter.SUBTITLE_FILL)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.font = Font(
                bold=row in (Formatter.QUOTE_ISSUER_ROW, Formatter.QUOTE_CUSTOMER_ROW),
                color=Formatter.MUTED_TEXT,
            )

        if logo:
            # ``get_column_letter`` rather than the cell's own attribute: the
            # title row is merged, and a merged cell carries no column letter.
            logo_column = get_column_letter(
                max(1, columns - Formatter.LOGO_COLUMN_INSET)
            )
            Formatter.anchor_logo(
                sheet, logo, f"{logo_column}{Formatter.QUOTE_ISSUER_ROW}"
            )

        status_row = Formatter.QUOTE_STATUS_ROW
        status_background, status_text = Formatter.STATUS_COLOURS.get(
            quote.status.value, (Formatter.SUBTITLE_FILL, Formatter.MUTED_TEXT)
        )
        sheet.merge_cells(f"A{status_row}:C{status_row}")
        translated = labels.get(f"status_{quote.status.value}", quote.status.value)
        status_cell = sheet.cell(
            row=status_row, column=1, value=f"{labels['status']}: {translated}"
        )
        status_cell.fill = PatternFill("solid", fgColor=status_background)
        status_cell.font = Font(bold=True, color=status_text)
        status_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False
        )
        for column in range(1, 4):
            sheet.cell(row=status_row, column=column).border = grid
        if quote.valid_until is not None:
            valid = sheet.cell(
                row=status_row,
                column=5,
                value=f"{labels['valid_until']}: {quote.valid_until}",
            )
            valid.font = Font(italic=True, color=Formatter.MUTED_TEXT)

        ordered = sorted(quote.lines, key=lambda line: (line.service_date, line.name))
        Formatter.write_quote_summary(sheet, quote, ordered, labels, language)

        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=column, value=header)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor=Formatter.HEADER_FILL)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = grid
        sheet.row_dimensions[header_row].height = 30

        band = PatternFill("solid", fgColor=Formatter.BAND_FILL)
        for offset, line in enumerate(ordered):
            row = header_row + 1 + offset
            sheet.cell(
                row=row, column=1, value=line.service_date
            ).number_format = Formatter.DATE_FORMAT
            sheet.cell(row=row, column=2, value=line.name)
            quantity = sheet.cell(row=row, column=3, value=line.duration_minutes / 60)
            quantity.number_format = Formatter.QUANTITY_FORMAT
            rate = sheet.cell(
                row=row,
                column=5,
                value=float(line.service_category.vat_rate()),
            )
            rate.number_format = Formatter.RATE_FORMAT
            for column, amount in enumerate((line.hourly_rate_ht,), start=4):
                cell = sheet.cell(
                    row=row, column=column, value=float(amount) if amount else None
                )
                cell.number_format = Formatter.MONEY_FORMAT
            for column, amount in enumerate(
                (line.total_ht, line.vat_amount, line.total_ttc), start=6
            ):
                cell = sheet.cell(
                    row=row, column=column, value=float(amount) if amount else None
                )
                cell.number_format = Formatter.MONEY_FORMAT
            for column in range(1, columns + 1):
                cell = sheet.cell(row=row, column=column)
                cell.border = grid
                if offset % 2:
                    cell.fill = band
                if column in (1, 3, 5):
                    cell.alignment = Alignment(horizontal="center")
                if column in (4, 6, 7, 8):
                    cell.alignment = Alignment(horizontal="right")

        total_row = Formatter.write_quote_totals(
            sheet, quote, ordered, header_row + 1 + len(ordered), columns, labels
        )
        Formatter.write_legal_terms(sheet, quote, total_row + 2, columns, labels)

        widths = [12, 34, 14, 20, 12, 18, 14, 18]
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[
                sheet.cell(row=header_row, column=column).column_letter
            ].width = width
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.sheet_properties.tabColor = Formatter.TITLE_FILL
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_title_rows = f"1:{header_row}"
        if ordered:
            sheet.auto_filter.ref = (
                f"A{header_row}:{last_column}{header_row + len(ordered)}"
            )

        stream = BytesIO()
        workbook.save(stream)
        payload = stream.getvalue()
        logger.info(
            "Rendered a %d-byte %s quote workbook for %s (%s), issued by %s.",
            len(payload),
            language.value,
            quote.reference,
            customer.id,
            company.name,
        )
        return payload
