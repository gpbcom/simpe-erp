from __future__ import annotations

# Standard library imports
from base64 import b64decode
from datetime import date, time
from decimal import Decimal
from email.message import EmailMessage
from io import BytesIO
from typing import List
from unittest.mock import AsyncMock, MagicMock

# Third-party imports
from openpyxl import load_workbook
import pytest

from models.auth.user import User

# First-party imports
from models.configuration.email_config import EmailConfig
from models.enums import (
    ServiceCategory,
    ContractType,
    InterventionStatus,
    Language,
    QuoteStatus,
    UserRole,
)
from models.companies.company import Company
from models.people.customer import Customer
from models.people.hca import Hca
from models.planning.hca_planning import HcaPlanning
from models.planning.intervention import Intervention
from models.quoting.quote import Quote
from models.quoting.quote_line import QuoteLine
from models.quoting.quote_type_week_aggregate import QuoteTypeWeekAggregate
from service.emails.emails import EmailService
from service.emails.exceptions import (
    MTEmailDeliveryFailed,
    MTEmailNoRecipient,
    MTEmailNotConfigured,
)
from service.utils.formatter import Formatter
from storage.s3.s3_storage import S3Storage

ADDRESS = {
    "street": "12 rue de Rivoli",
    "postal_code": "75004",
    "city": "Paris",
    "latitude": 48.8566,
    "longitude": 2.3522,
}


def _customer(customer_id: str = "customer-1") -> Customer:
    """Build a customer.

    Args:
        customer_id (str): The identifier to assign.

    Returns:
        Customer: The customer.
    """
    return Customer(
        id=customer_id,
        first_name="Marie",
        last_name="Durand",
        phone_number="+33612345678",
        email=f"{customer_id}@example.com",
        address=ADDRESS,
    )


def _hca(hca_id: str = "hca-1") -> Hca:
    """Build an assistant.

    Args:
        hca_id (str): The identifier to assign.

    Returns:
        Hca: The assistant.
    """
    return Hca(
        company_id="company-1",
        id=hca_id,
        first_name="Alice",
        last_name="Martin",
        phone_number="+33612345679",
        email=f"{hca_id}@example.com",
        address=ADDRESS,
        contract_type=ContractType.CDI,
    )


def _planning(interventions: int = 2) -> HcaPlanning:
    """Build a diary carrying a number of visits.

    Args:
        interventions (int): How many visits to place.

    Returns:
        HcaPlanning: The diary.
    """
    return HcaPlanning(
        hca_id="hca-1",
        hca_full_name="Alice Martin",
        period_start=date(2026, 8, 3),
        period_end=date(2026, 8, 9),
        interventions=[
            Intervention(
                name=f"Toilette {index}",
                intervention_type_id="type-1",
                company_id="company-1",
                service_category="necessity",
                quote_line_id=f"line-{index}",
                hca_id="hca-1",
                hca_full_name="Alice Martin",
                customer_id="customer-1",
                day=date(2026, 8, 5 - index),
                start_time=time(9 + index, 0),
                end_time=time(10 + index, 0),
                address=ADDRESS,
                status=InterventionStatus.PLANNED,
            )
            for index in range(interventions)
        ],
    )


def _manager(user_id: str = "user-1") -> User:
    """Build a manager account.

    Args:
        user_id (str): The identifier to assign.

    Returns:
        User: The account.
    """
    return User(
        company_id="company-1",
        id=user_id,
        email=f"{user_id}@example.com",
        full_name="Manager Account",
        role=UserRole.MANAGER,
    )


def _quote() -> Quote:
    """Build a priced, accepted quote.

    Returns:
        Quote: The quote.
    """
    return Quote(
        id="quote-1",
        reference="Q-2026-0001",
        customer_id="customer-1",
        company_id="company-1",
        status=QuoteStatus.ACCEPTED,
        lines=[
            QuoteLine(
                id="line-1",
                name="Toilette matin",
                intervention_type_id="type-1",
                service_category="necessity",
                service_date=date(2026, 8, 3),
                earliest_start=time(9, 0),
                latest_end=time(13, 0),
                duration_minutes=120,
                hourly_rate_ht=Decimal("31.91"),
                total_ht=Decimal("63.82"),
                vat_amount=Decimal("3.51"),
                total_ttc=Decimal("67.33"),
            )
        ],
        aggregates=[
            QuoteTypeWeekAggregate(
                intervention_type_id="type-1",
                service_category="necessity",
                intervention_type_name="Toilette",
                iso_year=2026,
                iso_week=32,
                week_start_date=date(2026, 8, 3),
                line_count=1,
                total_minutes=120,
                total_ht=Decimal("63.82"),
                vat_amount=Decimal("3.51"),
                total_ttc=Decimal("67.33"),
            )
        ],
    )


def _company() -> Company:
    """Build the agency issuing a quote.

    Returns:
        Company: The issuer named on the document.
    """
    return Company(
        id="company-1",
        name="Aide et Presence Paris",
        legal_form="SARL",
        share_capital=Decimal("10000"),
        registration_number="123456789",
        rcs_number="RCS Paris B 123 456 789",
        vat_number="FR12345678901",
        phone_number="01 23 45 67 89",
        contact_email="contact@simple-erp.fr",
        address={
            "street": "12 rue de Rivoli",
            "postal_code": "75004",
            "city": "Paris",
        },
    )


def _total_row(sheet, label: str) -> int:
    """Return the row a totals label was written on.

    Args:
        sheet: The rendered quote sheet.
        label (str): The label to find, such as "Total TTC".

    Returns:
        int: The row number.

    Notes:
        Searched rather than computed. The totals stack sits below a variable
        number of lines and one row per distinct VAT rate, so a literal row
        number in a test is a number that moves whenever a fixture gains a
        line — and fails describing the wrong thing when it does.

        The **last** match, not the first. The summary block above the table
        states the total including VAT as well, under the same word in
        French, and the first match is that one.
    """
    found = [
        cell.row for row in sheet.iter_rows() for cell in row if cell.value == label
    ]
    if not found:
        raise AssertionError(f"No totals row labelled {label!r} was written.")
    return found[-1]


class TestFormatter:
    """Tests for the Excel documents."""

    def test_a_planning_renders_a_readable_workbook(self) -> None:
        """The bytes open as a spreadsheet, not as a blob."""
        sheet = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        assert sheet.title == "Planning"
        assert sheet["A1"].value == "Planning — Alice Martin"
        assert [cell.value for cell in sheet[4]] == list(Formatter.PLANNING_HEADERS)

    def test_the_visits_are_written_in_the_order_they_happen(self) -> None:
        """An assistant works down the sheet, so it must be chronological.

        Notes:
            The fixture places its visits out of order deliberately: the second
            one happens the day before the first.
        """
        sheet = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        days = [sheet.cell(row=row, column=1).value.date() for row in (5, 6)]
        assert days == sorted(days)

    def test_an_empty_planning_says_so(self) -> None:
        """Nothing scheduled is a message, not a blank sheet."""
        sheet = load_workbook(
            BytesIO(Formatter.format_planning(_planning(interventions=0)))
        ).active
        assert sheet["A5"].value == "No intervention scheduled."

    def test_a_quote_renders_its_lines_and_total(self) -> None:
        """Every line and the total reach the sheet."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        assert sheet["A1"].value == "Devis Q-2026-0001"
        assert "Aide et Presence Paris" in sheet["A2"].value
        assert "Marie Durand" in sheet["A5"].value
        assert (
            sheet.cell(row=Formatter.QUOTE_HEADER_ROW + 1, column=2).value
            == "Toilette matin"
        )
        total = _total_row(sheet, "Total TTC")
        assert sheet.cell(row=total, column=8).value == pytest.approx(67.33)

    def test_an_unpriced_line_leaves_its_cells_empty(self) -> None:
        """Empty reads as "not priced yet"; a zero would read as "free"."""
        quote = _quote()
        quote.lines[0].total_ht = None
        quote.lines[0].total_ttc = None
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(quote, _customer(), _company()))
        ).active
        assert sheet.cell(row=Formatter.QUOTE_HEADER_ROW + 1, column=6).value is None
        assert sheet.cell(row=Formatter.QUOTE_HEADER_ROW + 1, column=8).value is None

    def test_amounts_are_numbers_not_strings(self) -> None:
        """A column of amounts must be summable by whoever opens it."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        cell = sheet.cell(row=Formatter.QUOTE_HEADER_ROW + 1, column=6)
        assert isinstance(cell.value, float)
        assert "€" in cell.number_format

    def test_every_method_is_static(self) -> None:
        """The formatter is a namespace, and is asked to stay one."""
        for name in ("format_planning", "format_quote"):
            attribute = Formatter.__dict__[name]
            assert isinstance(attribute, staticmethod)


class TestEmailService:
    """Tests for the delivery itself."""

    @pytest.fixture
    def sent(self, monkeypatch: pytest.MonkeyPatch) -> List[EmailMessage]:
        """Capture the messages instead of opening an SMTP connection.

        Args:
            monkeypatch (pytest.MonkeyPatch): Used to replace the delivery.

        Returns:
            List[EmailMessage]: The messages the service tried to send.
        """
        captured: List[EmailMessage] = []
        monkeypatch.setattr(
            EmailService, "_deliver", lambda self, message: captured.append(message)
        )
        monkeypatch.setenv("SMTP_USERNAME", "planner")
        monkeypatch.setenv("SMTP_PASSWORD", "secret")
        return captured

    @pytest.fixture
    def service(self) -> EmailService:
        """Return a service with outbound email switched on.

        Returns:
            EmailService: The service under test.
        """
        return EmailService(config=EmailConfig(enabled=True))

    async def test_a_planning_is_sent_to_the_assistant(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """The assistant's own stored address receives their diary."""
        await service.send_planning(_planning(), _hca())
        assert len(sent) == 1
        assert sent[0]["To"] == "hca-1@example.com"
        assert "planning" in sent[0]["Subject"].lower()

    async def test_a_planning_travels_as_a_spreadsheet(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """The document is attached, and it is a real workbook."""
        await service.send_planning(_planning(), _hca())
        attachment = next(sent[0].iter_attachments())
        assert attachment.get_filename().endswith(".xlsx")
        assert load_workbook(BytesIO(attachment.get_payload(decode=True))).active

    async def test_a_quote_is_sent_to_the_customer(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """The customer's own stored address receives their quote."""
        await service.send_quote(_quote(), _customer(), _company())
        assert sent[0]["To"] == "customer-1@example.com"
        assert "Q-2026-0001" in sent[0]["Subject"]

    async def test_a_disabled_mailbox_refuses_rather_than_pretends(
        self, sent: List[EmailMessage]
    ) -> None:
        """A planning nobody was sent must not look like one that was."""
        service = EmailService(config=EmailConfig(enabled=False))
        with pytest.raises(MTEmailNotConfigured):
            await service.send_planning(_planning(), _hca())
        assert sent == []

    async def test_missing_credentials_refuse_too(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Enabled but unconfigured is still not sendable."""
        monkeypatch.delenv("SMTP_USERNAME", raising=False)
        monkeypatch.delenv("SMTP_PASSWORD", raising=False)
        service = EmailService(config=EmailConfig(enabled=True))
        with pytest.raises(MTEmailNotConfigured):
            await service.send_quote(_quote(), _customer(), _company())

    async def test_an_smtp_failure_is_reported(
        self, service: EmailService, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """A refused conversation surfaces as the service's own exception."""
        monkeypatch.setenv("SMTP_USERNAME", "planner")
        monkeypatch.setenv("SMTP_PASSWORD", "secret")

        def _explode(self: EmailService, message: EmailMessage) -> None:
            raise OSError("connection refused")

        monkeypatch.setattr(EmailService, "_deliver", _explode)
        with pytest.raises(MTEmailDeliveryFailed):
            await service.send_planning(_planning(), _hca())

    async def test_one_bounced_assistant_does_not_stop_the_others(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """A run covers the workforce; one bad address must not cost the rest.

        Notes:
            The second assistant is built without an email address, which is
            what a bounced recipient looks like from here.
        """
        first, second = _hca("hca-1"), _hca("hca-2")
        plannings = [_planning(), _planning().model_copy(update={"hca_id": "hca-2"})]
        object.__setattr__(second, "email", "")
        delivered = await service.send_plannings(plannings, [first, second])
        assert delivered == 1
        assert len(sent) == 1

    async def test_a_planning_without_an_assistant_is_skipped(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """A diary whose assistant is gone has nobody to send to."""
        delivered = await service.send_plannings([_planning()], [])
        assert delivered == 0
        assert sent == []

    async def test_an_empty_recipient_is_refused(self, service: EmailService) -> None:
        """A message with nowhere to go is an error, not a silent no-op."""
        assistant = _hca()
        object.__setattr__(assistant, "email", "")
        with pytest.raises(MTEmailNoRecipient):
            await service.send_planning(_planning(), assistant)

    async def test_quotes_reach_their_own_customers(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """Each quote goes to the customer it names, and no other."""
        delivered = await service.send_quotes(
            [_quote()],
            [_customer("customer-9"), _customer("customer-1")],
            _company(),
        )
        assert delivered == 1
        assert sent[0]["To"] == "customer-1@example.com"


class TestTeamWorkbook:
    """Tests for the manager's consolidated document."""

    def test_each_assistant_gets_their_own_sheet(self) -> None:
        """A manager reads one round at a time; sheet tabs are the index."""
        plannings = [
            _planning().model_copy(
                update={"hca_id": "hca-1", "hca_full_name": "Alice Martin"}
            ),
            _planning().model_copy(
                update={"hca_id": "hca-2", "hca_full_name": "Bruno Petit"}
            ),
        ]
        workbook = load_workbook(BytesIO(Formatter.format_plannings(plannings)))
        assert workbook.sheetnames == ["Alice Martin", "Bruno Petit"]

    def test_the_sheets_are_ordered_by_name(self) -> None:
        """The same workbook two weeks running has its tabs in one order."""
        plannings = [
            _planning().model_copy(update={"hca_id": "b", "hca_full_name": "Zoe"}),
            _planning().model_copy(update={"hca_id": "a", "hca_full_name": "Ana"}),
        ]
        workbook = load_workbook(BytesIO(Formatter.format_plannings(plannings)))
        assert workbook.sheetnames == ["Ana", "Zoe"]

    def test_a_forbidden_character_does_not_break_the_workbook(self) -> None:
        """Excel refuses five characters in a sheet name; a name may carry them."""
        plannings = [
            _planning().model_copy(
                update={"hca_id": "hca-1", "hca_full_name": "Ana/Bea [test]"}
            )
        ]
        workbook = load_workbook(BytesIO(Formatter.format_plannings(plannings)))
        assert workbook.sheetnames == ["Ana Bea  test"]

    def test_two_assistants_sharing_a_name_keep_two_sheets(self) -> None:
        """A collision must not silently drop somebody's round."""
        plannings = [
            _planning().model_copy(update={"hca_id": "a", "hca_full_name": "Ana Roy"}),
            _planning().model_copy(update={"hca_id": "b", "hca_full_name": "Ana Roy"}),
        ]
        workbook = load_workbook(BytesIO(Formatter.format_plannings(plannings)))
        assert len(workbook.sheetnames) == 2

    def test_a_sheet_carries_that_assistants_visits(self) -> None:
        """Every sheet is the assistant's own week, laid out the same way."""
        plannings = [
            _planning().model_copy(
                update={"hca_id": "hca-1", "hca_full_name": "Alice Martin"}
            )
        ]
        sheet = load_workbook(BytesIO(Formatter.format_plannings(plannings)))[
            "Alice Martin"
        ]
        assert sheet["A1"].value == "Planning — Alice Martin"
        assert [cell.value for cell in sheet[4]] == list(Formatter.PLANNING_HEADERS)

    def test_the_week_is_named_on_every_sheet(self) -> None:
        """Both documents say which week they are, so neither can be misfiled."""
        single = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        assert single["A2"].value.startswith("Week 32 of 2026")


class TestWeeklyDispatch:
    """Tests for cutting a run's period into weeks."""

    @pytest.fixture
    def sent(self, monkeypatch: pytest.MonkeyPatch) -> List[EmailMessage]:
        """Capture the messages instead of opening an SMTP connection.

        Args:
            monkeypatch (pytest.MonkeyPatch): Used to replace the delivery.

        Returns:
            List[EmailMessage]: The messages the service tried to send.
        """
        captured: List[EmailMessage] = []
        monkeypatch.setattr(
            EmailService, "_deliver", lambda self, message: captured.append(message)
        )
        monkeypatch.setenv("SMTP_USERNAME", "planner")
        monkeypatch.setenv("SMTP_PASSWORD", "secret")
        return captured

    @pytest.fixture
    def service(self) -> EmailService:
        """Return a service with outbound email switched on.

        Returns:
            EmailService: The service under test.
        """
        return EmailService(config=EmailConfig(enabled=True))

    async def test_a_fortnight_is_sent_as_two_weeks(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """A run planned over two weeks produces two documents, not one.

        Notes:
            The period runs Monday 3 August to Sunday 16 August — two ISO
            weeks — with one visit in each.
        """
        planning = _planning(interventions=0).model_copy(
            update={
                "period_start": date(2026, 8, 3),
                "period_end": date(2026, 8, 16),
                "interventions": [
                    _planning()
                    .interventions[0]
                    .model_copy(update={"day": date(2026, 8, 4)}),
                    _planning()
                    .interventions[0]
                    .model_copy(update={"day": date(2026, 8, 11)}),
                ],
            }
        )
        delivered = await service.send_plannings([planning], [_hca()])
        assert delivered == 2
        subjects = sorted(message["Subject"] for message in sent)
        assert subjects == [
            "Your planning, 2026-08-03 to 2026-08-09",
            "Your planning, 2026-08-10 to 2026-08-16",
        ]

    async def test_each_week_carries_only_its_own_visits(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """A visit must appear in exactly one week's document."""
        planning = _planning(interventions=0).model_copy(
            update={
                "period_start": date(2026, 8, 3),
                "period_end": date(2026, 8, 16),
                "interventions": [
                    _planning()
                    .interventions[0]
                    .model_copy(update={"day": date(2026, 8, 4)}),
                    _planning()
                    .interventions[0]
                    .model_copy(update={"day": date(2026, 8, 11)}),
                ],
            }
        )
        await service.send_plannings([planning], [_hca()])
        rows = []
        for message in sent:
            attachment = next(message.iter_attachments())
            sheet = load_workbook(BytesIO(attachment.get_payload(decode=True))).active
            rows.append(sheet.cell(row=5, column=1).value.date())
        assert sorted(rows) == [date(2026, 8, 4), date(2026, 8, 11)]

    async def test_a_period_starting_midweek_still_starts_on_a_monday(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """Whatever the run was asked for, a document is a Monday-to-Sunday week."""
        planning = _planning().model_copy(
            update={"period_start": date(2026, 8, 5), "period_end": date(2026, 8, 7)}
        )
        await service.send_plannings([planning], [_hca()])
        assert sent[0]["Subject"] == "Your planning, 2026-08-03 to 2026-08-09"

    async def test_the_manager_receives_the_whole_workforce(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """One consolidated workbook per week, one sheet per assistant."""
        plannings = [
            _planning().model_copy(
                update={"hca_id": "hca-1", "hca_full_name": "Alice Martin"}
            ),
            _planning().model_copy(
                update={"hca_id": "hca-2", "hca_full_name": "Bruno Petit"}
            ),
        ]
        assistants = [_hca("hca-1"), _hca("hca-2")]
        delivered = await service.send_plannings(plannings, assistants, [_manager()])
        assert delivered == 3
        team = [m for m in sent if m["To"] == "user-1@example.com"]
        assert len(team) == 1
        attachment = next(team[0].iter_attachments())
        workbook = load_workbook(BytesIO(attachment.get_payload(decode=True)))
        assert workbook.sheetnames == ["Alice Martin", "Bruno Petit"]

    async def test_an_assistant_never_receives_the_team_workbook(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """The two documents are built from different inputs, not filtered.

        Notes:
            An assistant's copy comes from their own diary, so there is no
            arrangement in which a colleague's round can appear in it.
        """
        plannings = [
            _planning().model_copy(
                update={"hca_id": "hca-1", "hca_full_name": "Alice Martin"}
            ),
            _planning().model_copy(
                update={"hca_id": "hca-2", "hca_full_name": "Bruno Petit"}
            ),
        ]
        await service.send_plannings(
            plannings, [_hca("hca-1"), _hca("hca-2")], [_manager()]
        )
        for message in sent:
            if message["To"] == "user-1@example.com":
                continue
            attachment = next(message.iter_attachments())
            workbook = load_workbook(BytesIO(attachment.get_payload(decode=True)))
            assert workbook.sheetnames == ["Planning"]

    async def test_no_manager_means_no_team_email(
        self, service: EmailService, sent: List[EmailMessage]
    ) -> None:
        """The consolidated copy is sent to managers, or to nobody."""
        delivered = await service.send_plannings([_planning()], [_hca()])
        assert delivered == 1
        assert all(message["To"] == "hca-1@example.com" for message in sent)


class TestWorkbookStyling:
    """Tests for the documents looking like documents.

    Notes:
        These pin the *decisions*, not every colour: that a header is filled
        and legible, that a status is colour-coded, that rows alternate, that
        the header stays put when a long week is scrolled or printed. A
        workbook that quietly loses them still opens, still holds the right
        numbers, and looks like a dump.
    """

    def test_the_title_sits_in_a_filled_band(self) -> None:
        """The first thing a reader sees is a title, not a bare cell."""
        sheet = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        assert sheet["A1"].fill.fgColor.rgb == Formatter.TITLE_FILL
        assert sheet["A1"].font.bold is True
        assert sheet["A1"].font.color.rgb == "FFFFFFFF"

    def test_the_title_spans_the_table(self) -> None:
        """A band that stops after one column reads as a stray cell."""
        sheet = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        spans = {str(cells) for cells in sheet.merged_cells.ranges}
        assert "A1:G1" in spans

    def test_the_header_row_is_legible_on_its_fill(self) -> None:
        """White on navy, not black on navy."""
        sheet = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        header = sheet.cell(row=4, column=1)
        assert header.fill.fgColor.rgb == Formatter.HEADER_FILL
        assert header.font.color.rgb == "FFFFFFFF"
        assert header.font.bold is True

    @pytest.mark.parametrize(
        ("status", "expected"),
        [
            pytest.param(InterventionStatus.PLANNED, "FFDDEBF7", id="planned is blue"),
            pytest.param(
                InterventionStatus.CONFIRMED, "FFE2EFDA", id="confirmed is green"
            ),
            pytest.param(
                InterventionStatus.CANCELLED, "FFFCE4E4", id="cancelled is red"
            ),
        ],
    )
    def test_a_status_carries_its_own_colour(
        self, status: InterventionStatus, expected: str
    ) -> None:
        """A cancelled visit must not read like a confirmed one at a glance."""
        planning = _planning(interventions=1)
        planning.interventions[0].status = status
        sheet = load_workbook(BytesIO(Formatter.format_planning(planning))).active
        assert sheet.cell(row=5, column=7).fill.fgColor.rgb == expected

    def test_rows_alternate_so_a_long_week_stays_readable(self) -> None:
        """Banding is what keeps the eye on one row across seven columns."""
        sheet = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        assert sheet.cell(row=5, column=4).fill.fgColor.rgb != Formatter.BAND_FILL
        assert sheet.cell(row=6, column=4).fill.fgColor.rgb == Formatter.BAND_FILL

    def test_the_cells_are_ruled(self) -> None:
        """A table without lines is a list of words."""
        sheet = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        assert sheet.cell(row=5, column=1).border.left.style == "thin"

    def test_the_header_stays_put_when_scrolled_or_printed(self) -> None:
        """A week of visits runs past the fold on screen and on paper."""
        sheet = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        assert sheet.freeze_panes == "A5"
        assert sheet.print_title_rows == "$1:$4"
        assert sheet.page_setup.orientation == "landscape"

    def test_the_week_can_be_filtered(self) -> None:
        """A manager looking for one day should not have to read the rest."""
        sheet = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        assert sheet.auto_filter.ref == "A4:G6"

    def test_the_total_row_is_set_apart(self) -> None:
        """The figure a reader looks for last must not look like a visit."""
        sheet = load_workbook(BytesIO(Formatter.format_planning(_planning()))).active
        total = sheet.cell(row=7, column=1)
        assert total.value == "Total"
        assert total.fill.fgColor.rgb == Formatter.TOTAL_FILL
        assert total.border.top.style == "double"

    def test_every_sheet_of_the_team_workbook_is_styled(self) -> None:
        """The manager's copy is the same document, not a plainer one."""
        plannings = [
            _planning().model_copy(
                update={"hca_id": "hca-1", "hca_full_name": "Alice Martin"}
            )
        ]
        sheet = load_workbook(BytesIO(Formatter.format_plannings(plannings)))[
            "Alice Martin"
        ]
        assert sheet["A1"].fill.fgColor.rgb == Formatter.TITLE_FILL
        assert sheet.cell(row=4, column=1).fill.fgColor.rgb == Formatter.HEADER_FILL
        assert sheet.sheet_properties.tabColor.rgb == Formatter.TITLE_FILL

    def test_the_quote_status_is_a_colour_coded_pill(self) -> None:
        """Accepted, rejected and expired must be tellable apart at a glance."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        assert sheet["A7"].value == "Statut: accepté"
        assert sheet["A7"].fill.fgColor.rgb == "FFE2EFDA"

    def test_the_quote_status_pill_is_wide_enough_to_read(self) -> None:
        """Written into the date column alone it is clipped to "tatus: acce"."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        assert "A7:C7" in {str(cells) for cells in sheet.merged_cells.ranges}

    def test_amounts_are_right_aligned(self) -> None:
        """A column of money that does not line up cannot be scanned."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        assert (
            sheet.cell(
                row=Formatter.QUOTE_HEADER_ROW + 1, column=8
            ).alignment.horizontal
            == "right"
        )

    def test_the_quote_total_is_emphasised(self) -> None:
        """The figure the customer is being asked to pay is the point."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        total = sheet.cell(row=_total_row(sheet, "Total TTC"), column=8)
        assert total.font.bold is True
        assert total.fill.fgColor.rgb == Formatter.TOTAL_FILL


class TestQuoteLanguage:
    """Tests for the language a quote and its covering note are written in."""

    def test_french_is_the_default(self) -> None:
        """A caller that names no language gets French.

        Notes:
            French rather than the caller's browser: this is a French agency,
            and a quote reaching a customer in English because nobody set a
            preference is the wrong failure to default into.
        """
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active

        assert sheet["A1"].value.startswith("Devis")
        assert sheet.title == "Devis"

    @pytest.mark.parametrize(
        ("language", "title", "service_header"),
        [
            pytest.param(Language.FR, "Devis", "Prestation", id="French"),
            pytest.param(Language.EN, "Quote", "Service", id="English"),
        ],
    )
    def test_the_whole_document_follows_the_language(
        self, language: Language, title: str, service_header: str
    ) -> None:
        """Title, sheet name and column headings all move together.

        Args:
            language (Language): The language asked for.
            title (str): The word the title should start with.
            service_header (str): The heading over the service column.

        Notes:
            Asserted across three places rather than one. A catalogue wired
            into the title and forgotten in the headers produces a document
            that is half-translated, which reads worse than one that is not
            translated at all.
        """
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company(), language))
        ).active

        assert sheet["A1"].value.startswith(title)
        assert sheet.title == title
        assert (
            sheet.cell(row=Formatter.QUOTE_HEADER_ROW, column=2).value == service_header
        )

    def test_the_status_word_is_translated(self) -> None:
        """The pill carries a word, so the word has to be in the language."""
        french = load_workbook(
            BytesIO(
                Formatter.format_quote(_quote(), _customer(), _company(), Language.FR)
            )
        ).active
        english = load_workbook(
            BytesIO(
                Formatter.format_quote(_quote(), _customer(), _company(), Language.EN)
            )
        ).active

        assert french["A7"].value == "Statut: accepté"
        assert english["A7"].value == "Status: accepted"

    @pytest.mark.parametrize(
        ("language", "expected"),
        [
            pytest.param(Language.FR, "2 h 00", id="French"),
            pytest.param(Language.EN, "2h 00m", id="English"),
        ],
    )
    def test_a_duration_is_written_out_not_left_in_minutes(
        self, language: Language, expected: str
    ) -> None:
        """ "1110" is the solver's unit, not an answer for a customer.

        Args:
            language (Language): The language asked for.
            expected (str): How the duration should read.
        """
        assert Formatter.format_duration(120, language) == expected


class TestQuoteParties:
    """Tests for the two parties a quote has to name."""

    def test_the_issuing_agency_is_named(self) -> None:
        """**A quote naming only the recipient cannot be acted on.**

        Notes:
            The customer needs to know who is offering, under what registered
            number, and where to reply. None of that was on the document
            before: a customer carries no company and neither does a quote, so
            the agency comes from the account that issued it.
        """
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active

        assert "Aide et Presence Paris" in sheet["A2"].value
        details = sheet["A3"].value
        assert "12 rue de Rivoli" in details
        assert "123456789" in details
        assert "contact@simple-erp.fr" in details

    def test_the_customer_is_named_with_their_address(self) -> None:
        """The recipient, and where the care is delivered."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active

        assert "Marie Durand" in sheet["A5"].value
        assert "rue" in sheet["A6"].value.lower()

    def test_an_agency_with_no_registration_number_prints_no_label(self) -> None:
        """An optional field left unset is not a data error to display.

        Notes:
            Joining only what is set. "SIRET None" beside a real address reads
            as a fault on a document somebody is being asked to sign.
        """
        bare = Company(id="company-1", name="Petite Agence")
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), bare))
        ).active

        assert not sheet["A3"].value
        assert "Petite Agence" in sheet["A2"].value


class TestQuoteSummary:
    """Tests for the four figures asked of a quote before its price."""

    def test_the_period_and_duration_are_stated(self) -> None:
        """When does this start, when does it end, and how much care is it."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        row = Formatter.QUOTE_SUMMARY_ROW

        assert sheet.cell(row=row, column=1).value == "Période"
        assert "2026-08-03" in str(sheet.cell(row=row, column=3).value)
        assert sheet.cell(row=row, column=6).value == "Interventions"
        assert sheet.cell(row=row, column=7).value == 1
        assert sheet.cell(row=row + 1, column=1).value == "Durée totale"
        assert sheet.cell(row=row + 1, column=3).value == "2 h 00"

    def test_the_summary_total_matches_the_table_total(self) -> None:
        """Two figures for one number is one figure too many.

        Notes:
            The summary states the total incl. VAT and so does the table's last
            row. They are computed from the same method, and this is what says
            so — a summary carrying a stale figure is worse than no summary.
        """
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        summary = sheet.cell(row=Formatter.QUOTE_SUMMARY_ROW + 1, column=7).value
        table = sheet.cell(row=_total_row(sheet, "Total TTC"), column=8).value

        assert summary == pytest.approx(table)

    def test_the_summary_is_translated(self) -> None:
        """Its labels come from the same catalogue as everything else."""
        sheet = load_workbook(
            BytesIO(
                Formatter.format_quote(_quote(), _customer(), _company(), Language.EN)
            )
        ).active
        row = Formatter.QUOTE_SUMMARY_ROW

        assert sheet.cell(row=row, column=1).value == "Period"
        assert sheet.cell(row=row + 1, column=1).value == "Total duration"

    def test_every_line_carries_its_duration_and_both_prices(self) -> None:
        """Duration, price excluding VAT and price including it, per line."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        row = Formatter.QUOTE_HEADER_ROW + 1

        assert sheet.cell(row=row, column=3).value == pytest.approx(2.0)
        assert sheet.cell(row=row, column=4).value == pytest.approx(31.91)
        assert sheet.cell(row=row, column=6).value == pytest.approx(63.82)
        assert sheet.cell(row=row, column=8).value == pytest.approx(67.33)


class TestQuoteLegalForm:
    """Tests for the mentions a quote must carry to be a quote at all."""

    def test_the_issuer_is_named_with_its_legal_form(self) -> None:
        """**"Aide et Presence Paris" is a trading name, not a legal person.**

        Notes:
            The customer is contracting with the SARL, and the legal form is
            part of how a French company must identify itself on a commercial
            document.
        """
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active

        assert "Aide et Presence Paris SARL" in sheet["A2"].value

    def test_an_agency_with_no_legal_form_prints_its_name_alone(self) -> None:
        """A field not filled in yet must not take the quote away.

        Notes:
            Every legal column is nullable, because none has a safe default.
            An agency fills them in on its own screen; until it does, the
            document says what it can.
        """
        bare = Company(id="company-1", name="Petite Agence")
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), bare))
        ).active

        assert "Petite Agence" in sheet["A2"].value
        assert "None" not in sheet["A2"].value

    def test_the_issuer_line_carries_every_legal_identifier(self) -> None:
        """SIRET, RCS, VAT number, capital, phone and email."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        details = sheet["A3"].value

        for fragment in (
            "12 rue de Rivoli",
            "123456789",
            "RCS Paris B 123 456 789",
            "FR12345678901",
            "01 23 45 67 89",
            "contact@simple-erp.fr",
        ):
            assert fragment in details, f"{fragment!r} is missing from the issuer."

    def test_the_capital_is_labelled_and_carries_its_currency(self) -> None:
        """ "Capital social 10 000,00 €", not a bare number."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active

        assert "Capital social" in sheet["A3"].value
        assert "€" in sheet["A3"].value


class TestQuoteTotalsBlock:
    """Tests for the untaxed / tax / total stack the document ends on."""

    def test_the_totals_are_summed_from_the_lines_printed_above(self) -> None:
        """**A total that disagrees with its own column is the worst failure.**

        Notes:
            The quote's own ``total_ht`` sums the *aggregates*, which the
            pricing service computes. A quote reaching the renderer without
            them would print zero under a column of real amounts — and look
            correct doing it.
        """
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        untaxed = sheet.cell(row=_total_row(sheet, "Total HT"), column=8).value
        total = sheet.cell(row=_total_row(sheet, "Total TTC"), column=8).value

        assert untaxed == pytest.approx(63.82)
        assert total == pytest.approx(67.33)

    def test_the_tax_is_broken_down_per_rate(self) -> None:
        """Two rates, two lines — not one "VAT" figure.

        Notes:
            Home care is billed at 5.5% for a necessity and 20% for a comfort
            service. A single tax line gives the customer no way to check
            either, and an accountant no way to post it.
        """
        quote = _quote()
        comfort = quote.lines[0].model_copy(
            update={
                "id": "line-2",
                "service_category": ServiceCategory.COMFORT,
                "total_ht": Decimal("200.00"),
                "vat_amount": Decimal("40.00"),
                "total_ttc": Decimal("240.00"),
            }
        )
        quote.lines.append(comfort)
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(quote, _customer(), _company()))
        ).active

        assert sheet.cell(row=_total_row(sheet, "TVA 5.5%"), column=8).value == (
            pytest.approx(3.51)
        )
        assert sheet.cell(row=_total_row(sheet, "TVA 20%"), column=8).value == (
            pytest.approx(40.00)
        )

    @pytest.mark.parametrize(
        ("rate", "expected"),
        [
            pytest.param(Decimal("0.055"), "5.5%", id="a fractional rate"),
            pytest.param(Decimal("0.20"), "20%", id="a whole rate"),
            pytest.param(Decimal("0.2"), "20%", id="the same, written shorter"),
        ],
    )
    def test_a_rate_drops_its_trailing_zeros(
        self, rate: Decimal, expected: str
    ) -> None:
        """ "TVA 5.500%" reads as a rate somebody mistyped.

        Args:
            rate (Decimal): The rate as a fraction.
            expected (str): How it should read.
        """
        assert Formatter.format_rate(rate) == expected

    def test_the_line_quantity_is_hours_not_minutes(self) -> None:
        """**Otherwise the arithmetic on the page is wrong by sixty.**

        Notes:
            The unit price beside it is an hourly rate, so a customer checking
            quantity × price against the amount has to be able to get the
            amount.
        """
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        row = Formatter.QUOTE_HEADER_ROW + 1
        quantity = sheet.cell(row=row, column=3).value
        unit = sheet.cell(row=row, column=4).value
        amount = sheet.cell(row=row, column=6).value

        assert quantity == pytest.approx(2.0)
        assert quantity * unit == pytest.approx(amount, abs=0.01)


class TestQuoteLegalTerms:
    """Tests for the footer that makes the document an offer."""

    @pytest.mark.parametrize(
        ("language", "fragment"),
        [
            pytest.param(Language.FR, "Bon pour accord", id="French"),
            pytest.param(Language.EN, "Bon pour accord", id="English"),
        ],
    )
    def test_the_acceptance_mention_is_present(
        self, language: Language, fragment: str
    ) -> None:
        """**A quote without it is a price list, not an offer.**

        Args:
            language (Language): The language asked for.
            fragment (str): The wording that must appear.

        Notes:
            French law requires a commercial offer to be returned signed before
            the work begins. The obligation is French whatever language the
            reader chose, so the mention appears in both — the language decides
            the words, not the law.
        """
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company(), language))
        ).active
        text = " ".join(
            str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value
        )

        assert fragment in text

    def test_the_terms_state_payment_and_penalties(self) -> None:
        """When payment falls due, and what happens if it does not."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        text = " ".join(
            str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value
        )

        assert "30 jours" in text
        assert "L441-10" in text
        assert _quote().reference in text

    def test_a_quote_with_no_expiry_states_no_validity(self) -> None:
        """An offer promising to stand "until None" is worse than none.

        Notes:
            A draft carries no ``valid_until`` yet. The sentence is omitted
            rather than printed with a blank in it.
        """
        quote = _quote()
        quote.valid_until = None
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(quote, _customer(), _company()))
        ).active
        text = " ".join(
            str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value
        )

        assert "valable" not in text.lower()
        assert "None" not in text

    def test_the_customer_gets_somewhere_to_sign(self) -> None:
        """The signature block is what turns it into an agreement."""
        sheet = load_workbook(
            BytesIO(Formatter.format_quote(_quote(), _customer(), _company()))
        ).active
        text = " ".join(
            str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value
        )

        assert "signature du client" in text.lower()


#: A one-pixel PNG, the smallest thing Pillow will open. Embedded rather than
#: generated so the test asserts on the renderer, not on an image library.
ONE_PIXEL_PNG = b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAE"
    "hQGAhKmMIQAAAABJRU5ErkJggg=="
)


class TestTheQuoteCarriesTheAgencyLogo:
    """Tests for the letterhead on the document a customer signs."""

    def test_a_logo_is_drawn_on_the_sheet(self) -> None:
        """The image travels in the workbook, not as a link.

        Notes:
            A link would resolve only for a reader who could reach the object
            store, which a customer cannot — the document has to carry the
            picture itself.
        """
        payload = Formatter.format_quote(
            _quote(), _customer(), _company(), logo=ONE_PIXEL_PNG
        )

        sheet = load_workbook(BytesIO(payload)).active

        assert len(sheet._images) == 1

    def test_no_logo_leaves_the_sheet_as_it_was(self) -> None:
        """An agency that has not uploaded one prints what it always did."""
        payload = Formatter.format_quote(_quote(), _customer(), _company())

        sheet = load_workbook(BytesIO(payload)).active

        assert sheet._images == []

    def test_a_corrupt_logo_does_not_stop_the_quote(self) -> None:
        """**A quote must go out even when its letterhead is unreadable.**

        Notes:
            Refusing to produce a priced offer because a decoration could not
            be decoded would turn a cosmetic problem into a commercial one.
        """
        payload = Formatter.format_quote(
            _quote(), _customer(), _company(), logo=b"not an image at all"
        )

        sheet = load_workbook(BytesIO(payload)).active

        assert sheet._images == []
        assert sheet["A1"].value is not None

    def test_the_logo_is_scaled_down_but_never_up(self) -> None:
        """A small mark was meant to be small.

        Notes:
            Enlarging it to fill the space would print it blurred on a document
            the agency is identified by.
        """
        payload = Formatter.format_quote(
            _quote(), _customer(), _company(), logo=ONE_PIXEL_PNG
        )

        drawn = load_workbook(BytesIO(payload)).active._images[0]

        assert drawn.width == 1
        assert drawn.height == 1


class TestFetchingTheLogoForAQuote:
    """Tests for how the mailer obtains the image it embeds."""

    async def test_an_agency_with_no_logo_fetches_nothing(self) -> None:
        """No logo is not a reason to talk to the object store."""
        logos = MagicMock(spec=S3Storage)
        logos.fetch_logo = AsyncMock(return_value=ONE_PIXEL_PNG)
        service = EmailService(config=EmailConfig(enabled=True), logos=logos)

        assert await service._fetch_logo(_company()) is None
        logos.fetch_logo.assert_not_awaited()

    async def test_a_deployment_with_no_object_store_fetches_nothing(self) -> None:
        """Quotes still go out; they simply have no letterhead."""
        service = EmailService(config=EmailConfig(enabled=True))

        agency = _company().model_copy(
            update={
                "logo_url": (
                    "https://minio.internal/simple-erp/company-logos/company-1/a.png"
                )
            }
        )

        assert await service._fetch_logo(agency) is None

    async def test_an_unreadable_logo_is_reported_not_raised(self) -> None:
        """**The failure mode that must not reach the customer.**

        Notes:
            The store already reports rather than raises; this asserts the
            mailer does not turn that ``None`` back into an exception on the
            way to the renderer.
        """
        logos = MagicMock(spec=S3Storage)
        logos.fetch_logo = AsyncMock(return_value=None)
        service = EmailService(config=EmailConfig(enabled=True), logos=logos)

        agency = _company().model_copy(
            update={
                "logo_url": (
                    "https://minio.internal/simple-erp/company-logos/company-1/a.png"
                )
            }
        )

        assert await service._fetch_logo(agency) is None
